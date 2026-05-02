#!/usr/bin/env python3
from __future__ import annotations

import io
import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from PIL import Image, ImageOps


ROOT = Path(__file__).resolve().parent.parent
MAPPING_JSON = Path(r"C:\Users\Administrator\Downloads\akel-image-mapping-updates.json")
SOURCE_DIR = Path("D:/AkelWebSite") / "\u0130MAJ FOTO\u011eRAFLAR"
EXCEL_PATH = ROOT / "akel-urunler.xlsx"
WEBSITE = ROOT / "akel-melamin-design-system" / "project" / "ui_kits" / "website"
IMG_DIR = WEBSITE / "img"
JSON_PATH = WEBSITE / "data" / "products.json"
BACKUP_DIR = ROOT / "product-image-matcher" / "site-image-backup"

MAX_SIZE = (1600, 1600)
WEBP_QUALITY = 82

COLS = [
    "sku",
    "kategori",
    "ad_tr",
    "ad_en",
    "renk",
    "foto_dosya_adi",
    "boyut",
    "cap_mm",
    "en_mm",
    "boy_mm",
    "yukseklik_mm",
    "agirlik_g",
    "koli_adet",
    "koli_cm",
    "koli_agirlik_kg",
    "aciklama_tr",
    "aciklama_en",
    "kullanim_alani",
    "barkod",
    "aktif",
    "oncelik",
    "ozel_logolu_uygun",
    "etiket",
]


def selected_file(change: dict) -> str:
    return change.get("secilen_imaj_dosyasi") or change.get("yeni_foto_dosya_adi") or ""


def load_changes() -> list[dict]:
    payload = json.loads(MAPPING_JSON.read_text(encoding="utf-8"))
    changes = payload.get("changes")
    if not isinstance(changes, list) or not changes:
        raise ValueError("Mapping JSON has no changes.")
    return changes


def webp_bytes(source_name: str) -> bytes:
    source_path = SOURCE_DIR / source_name
    if not source_path.exists():
        raise FileNotFoundError(source_path)
    with Image.open(source_path) as image:
        image = ImageOps.exif_transpose(image)
        image.thumbnail(MAX_SIZE, Image.Resampling.LANCZOS)
        if image.mode not in ("RGB", "RGBA"):
            image = image.convert("RGB")
        out = io.BytesIO()
        image.save(out, "WEBP", quality=WEBP_QUALITY, method=6)
        return out.getvalue()


def backup_existing(path: Path) -> None:
    if not path.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / path.name
    if not backup_path.exists():
        shutil.copy2(path, backup_path)


def preserve_plain_image(product_image: Path, plain_image: Path) -> bool:
    if plain_image.exists():
        return False

    backup_path = BACKUP_DIR / product_image.name
    source = backup_path if backup_path.exists() else product_image
    if not source.exists():
        return False

    shutil.copy2(source, plain_image)
    return True


def write_product_webps(changes: list[dict]) -> int:
    IMG_DIR.mkdir(parents=True, exist_ok=True)
    source_cache: dict[str, bytes] = {}
    written = 0
    plain_written = 0

    for change in changes:
        source_name = selected_file(change)
        old_base = str(change.get("eski_foto_dosya_adi") or "").strip()
        if not source_name or not old_base:
            raise ValueError(f"Invalid mapping row: {change}")
        if source_name not in source_cache:
            source_cache[source_name] = webp_bytes(source_name)

        dest = IMG_DIR / f"{old_base}.webp"
        plain_dest = IMG_DIR / f"sade-{old_base}.webp"
        backup_existing(dest)
        if preserve_plain_image(dest, plain_dest):
            plain_written += 1
        dest.write_bytes(source_cache[source_name])
        written += 1

    return written, plain_written


def change_lookup(changes: list[dict]) -> dict[tuple[str, str], str]:
    # Keyed by SKU + category so we can restore product image names even if a
    # previous attempt changed foto_dosya_adi to a shared source-image name.
    lookup: dict[tuple[str, str], str] = {}
    for change in changes:
        key = (str(change.get("sku") or "").strip(), str(change.get("kategori") or "").strip())
        old_base = str(change.get("eski_foto_dosya_adi") or "").strip()
        if key in lookup and lookup[key] != old_base:
            raise ValueError(f"Ambiguous mapping for {key}")
        lookup[key] = old_base
    return lookup


def update_excel(changes: list[dict]) -> int:
    lookup = change_lookup(changes)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    if "\u00dcr\u00fcnler" not in wb.sheetnames:
        raise ValueError("Excel sheet not found: \u00dcr\u00fcnler")
    ws = wb["\u00dcr\u00fcnler"]
    foto_col = COLS.index("foto_dosya_adi") + 1
    updated = 0

    for row_index in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row_index, column=COLS.index("sku") + 1).value or "").strip()
        kategori = str(ws.cell(row=row_index, column=COLS.index("kategori") + 1).value or "").strip()
        key = (sku, kategori)
        if key not in lookup:
            continue
        current = str(ws.cell(row=row_index, column=foto_col).value or "").strip()
        desired = lookup[key]
        if current != desired:
            ws.cell(row=row_index, column=foto_col).value = desired
        updated += 1

    wb.save(EXCEL_PATH)
    return updated


def update_products_json(changes: list[dict]) -> int:
    lookup = change_lookup(changes)
    payload = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    updated = 0
    for product in payload.get("products", []):
        key = (str(product.get("sku") or ""), str(product.get("kategori") or ""))
        if key not in lookup:
            continue
        product["foto_dosya_adi"] = lookup[key]
        plain_base = f"sade-{lookup[key]}"
        if (IMG_DIR / f"{plain_base}.webp").exists():
            product["sade_foto_dosya_adi"] = plain_base
        updated += 1
    JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return updated


def main() -> int:
    if not MAPPING_JSON.exists():
        raise FileNotFoundError(MAPPING_JSON)
    if not SOURCE_DIR.exists():
        raise FileNotFoundError(SOURCE_DIR)

    changes = load_changes()
    written, plain_written = write_product_webps(changes)
    excel_updated = update_excel(changes)
    json_updated = update_products_json(changes)

    print(f"changes={len(changes)}")
    print(f"product_webps_written={written}")
    print(f"plain_webps_written={plain_written}")
    print(f"excel_rows_matched={excel_updated}")
    print(f"json_products_matched={json_updated}")
    print(f"backup_dir={BACKUP_DIR.relative_to(ROOT)}")
    print(f"finished_at={datetime.now().isoformat(timespec='seconds')}")

    if written != len(changes) or json_updated != len(changes):
        raise SystemExit("Not all JSON mappings were applied.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
