#!/usr/bin/env python3
"""
Akel Melamin — Excel → JSON dönüştürücü
Kullanıcı akel-urunler.xlsx dosyasını doldurur,
bu script o dosyayı okuyup sitenin okuduğu products.json'u yeniden üretir.

Kullanım:
    python3 excel-to-products.py

Çıktı:
    ui_kits/website/data/products.json (overwrite)

Validasyon:
    - Zorunlu alanlar (sku, kategori, ad_tr, renk, foto_dosya_adi)
    - Kategori 13 listesinden mi?
    - Foto dosyası /img/ altında gerçekten var mı?
    - SKU tekilliği
"""
from __future__ import annotations
import json
import sys
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("HATA: openpyxl yüklü değil.  pip install openpyxl")
    sys.exit(1)

# ---- Paths ----
ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / "akel-urunler.xlsx"
SITE = ROOT / "akel-melamin-design-system" / "project" / "ui_kits" / "website"
IMG_DIR = SITE / "img"
JSON_OUT = SITE / "data" / "products.json"

# ---- Kategori / renk whitelist'leri ----
VALID_CATEGORIES = {
    "kare",
    "kare-siyah",
    "international",
    "elegant",
    "acikbufe",
    "tepsiler",
    "eco",
    "diplomat",
    "gunluk",
    "masaustu",
    "kids",
    "bolmeli",
    "logolu",
}
VALID_COLORS = {"beyaz", "siyah", "desenli", "kirmizi", "mavi", "yesil"}
REQUIRED = ["sku", "kategori", "ad_tr", "renk", "foto_dosya_adi"]

CATEGORIES = [
    {"key": "kare", "tr": "Kare Seri", "en": "Square Series"},
    {"key": "kare-siyah", "tr": "Kare Seri Siyah", "en": "Square Series · Black"},
    {"key": "international", "tr": "International Seri", "en": "International Series"},
    {"key": "masaustu", "tr": "Masaüstü Aksesuarları", "en": "Tabletop Accessories"},
    {"key": "acikbufe", "tr": "Açık Büfe · Servis", "en": "Buffet · Serving"},
    {"key": "elegant", "tr": "Elegant Seri", "en": "Elegant Series"},
    {"key": "kids", "tr": "Çocuk Tabakları", "en": "Kids' Plates"},
    {"key": "bolmeli", "tr": "Bölmeli Tabaklar", "en": "Compartment Plates"},
    {"key": "tepsiler", "tr": "Tepsiler", "en": "Trays"},
    {"key": "eco", "tr": "Eco Seri", "en": "Eco Series"},
    {"key": "diplomat", "tr": "Diplomat Kayık Seri", "en": "Diplomat Boat Series"},
    {"key": "gunluk", "tr": "Günlük Desenli Seri", "en": "Everyday Pattern Line"},
    {"key": "logolu", "tr": "Özel Logolu Ürünler", "en": "Custom-logo Products"},
]

# Excel column key -> JSON key (1:1 in our case, but keep explicit)
COLS = [
    "sku",
    "kategori",
    "ad_tr",
    "ad_en",
    "renk",
    "foto_dosya_adi",
    "boyut",
    "cap_mm",
    "en_mm",
    "boy_mm",
    "yukseklik_mm",
    "agirlik_g",
    "koli_adet",
    "koli_cm",
    "koli_agirlik_kg",
    "aciklama_tr",
    "aciklama_en",
    "kullanim_alani",
    "barkod",
    "aktif",
    "oncelik",
    "ozel_logolu_uygun",
    "etiket",
]


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def is_row_empty(values) -> bool:
    return all(norm(v) in (None, "") for v in values)


def main():
    if not EXCEL.exists():
        print(f"HATA: Excel bulunamadı: {EXCEL}")
        return 1

    print(f"Okunuyor: {EXCEL.name}")
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    if "Ürünler" not in wb.sheetnames:
        print("HATA: 'Ürünler' sayfası bulunamadı.")
        return 1
    ws = wb["Ürünler"]

    products = []
    errors = []
    warnings = []
    seen_sku = set()

    # Row 1 = header, row 2 = example (italic), data starts row 3 onwards.
    # But user might also start at row 2 if they delete the sample. So we try
    # each row and only keep ones with SKU filled AND not labeled as example.
    for r in range(2, ws.max_row + 1):
        raw = [ws.cell(row=r, column=i + 1).value for i in range(len(COLS))]
        if is_row_empty(raw):
            continue

        row = dict(zip(COLS, [norm(v) for v in raw]))

        # Skip the seeded italic example row if user didn't overwrite it
        if (
            row.get("sku") == "AKL-KARE-B-24"
            and row.get("ad_tr") == "Kare Düz Tabak 24 cm"
        ):
            # Heuristic: if font is italic (example marker), skip
            cell = ws.cell(row=r, column=1)
            if cell.font and cell.font.italic:
                continue

        # --- Validate required ---
        missing = [k for k in REQUIRED if not row.get(k)]
        if missing:
            errors.append(f"Satır {r}: eksik zorunlu alan → {', '.join(missing)}")
            continue

        # --- Validate kategori ---
        if row["kategori"] not in VALID_CATEGORIES:
            errors.append(
                f"Satır {r} ({row['sku']}): geçersiz kategori '{row['kategori']}'"
            )
            continue

        # --- Validate renk ---
        if row["renk"] not in VALID_COLORS:
            warnings.append(
                f"Satır {r} ({row['sku']}): beklenmeyen renk '{row['renk']}'"
            )

        # --- SKU tekilliği ---
        if row["sku"] in seen_sku:
            errors.append(f"Satır {r}: SKU tekrarı '{row['sku']}'")
            continue
        seen_sku.add(row["sku"])

        # --- Foto var mı? ---
        photo_name = row["foto_dosya_adi"]
        matches = list(IMG_DIR.glob(f"{photo_name}.*"))
        if not matches:
            warnings.append(
                f"Satır {r} ({row['sku']}): foto dosyası bulunamadı → "
                f"{photo_name}.jpg|png|webp  ({IMG_DIR})"
            )

        # --- Type coercions ---
        for num_key in (
            "cap_mm",
            "en_mm",
            "boy_mm",
            "yukseklik_mm",
            "agirlik_g",
            "koli_adet",
            "oncelik",
        ):
            if row.get(num_key) is not None:
                try:
                    row[num_key] = int(float(row[num_key]))
                except (ValueError, TypeError):
                    warnings.append(f"Satır {r} ({row['sku']}): {num_key} sayı değil")
                    row[num_key] = None
        if row.get("koli_agirlik_kg") is not None:
            try:
                row["koli_agirlik_kg"] = float(row["koli_agirlik_kg"])
            except (ValueError, TypeError):
                row["koli_agirlik_kg"] = None

        # Default aktif=evet if blank
        if not row.get("aktif"):
            row["aktif"] = "evet"

        products.append(row)

    # ---- Sort: aktif önce, oncelik'e göre (boş = sonda), sonra SKU ----
    def sort_key(p):
        return (
            p.get("aktif") != "evet",
            p.get("oncelik") if p.get("oncelik") is not None else 999,
            p.get("kategori") or "",
            p.get("sku") or "",
        )

    products.sort(key=sort_key)

    # ---- Write JSON ----
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    out = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source": "excel",
        "categories": CATEGORIES,
        "products": products,
    }
    JSON_OUT.write_text(
        json.dumps(out, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    # ---- Report ----
    print()
    print(f"✓ {len(products)} ürün yazıldı → {JSON_OUT}")
    if warnings:
        print()
        print(f"⚠  {len(warnings)} UYARI (kritik değil):")
        for w in warnings:
            print(f"   · {w}")
    if errors:
        print()
        print(f"✗  {len(errors)} HATA (bu ürünler atlandı):")
        for e in errors:
            print(f"   · {e}")
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
